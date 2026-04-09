"""
CORE – Excel til JavaScript-konverter
──────────────────────────────────────
Kjør dette skriptet etter at du har redigert database.xlsx:

    python3 convert.py

Skriptet leser database.xlsx og oppdaterer cities.js og experiences.js
automatisk. Du trenger ikke røre JavaScript-filene manuelt.
"""

import sys
import json
from pathlib import Path
from openpyxl import load_workbook

BASE     = Path(__file__).parent
DB_FILE  = BASE / "database.xlsx"
CITIES_JS = BASE / "cities.js"
EXP_JS   = BASE / "experiences.js"

VALID_SELSKAP = {"solo", "par", "venner", "familie"}
VALID_KOSTNAD = {"gratis", "$", "$$", "$$$"}
VALID_TID     = {"morgen", "formiddag", "ettermiddag", "kveld"}

def split_comma(val):
    return [v.strip() for v in str(val).split(",") if v.strip()]

def sheet_to_dicts(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        obj = {headers[i]: (row[i] if row[i] is not None else "") for i in range(len(headers))}
        result.append(obj)
    return result

if not DB_FILE.exists():
    print("❌  Fant ikke database.xlsx i samme mappe som convert.py")
    sys.exit(1)

wb = load_workbook(DB_FILE, data_only=True)

# ── Byer ───────────────────────────────────────────────────────
if "Byer" not in wb.sheetnames:
    print('❌  Fant ikke arket "Byer" i database.xlsx')
    sys.exit(1)

city_rows = sheet_to_dicts(wb["Byer"])
cities = []
for r in city_rows:
    name = str(r.get("name", "")).strip()
    if not name:
        continue
    cities.append({
        "name":  name,
        "image": str(r.get("image", "")).strip(),
        "alt":   str(r.get("alt", "")).strip(),
    })

cities_js = f"""/*
 * ============================================================
 *  CORE – BY-DATABASE  (auto-generert av convert.py)
 * ============================================================
 *  IKKE rediger denne filen direkte.
 *  Gjør endringer i database.xlsx og kjør:  python3 convert.py
 * ============================================================
 */

const CITIES = {json.dumps(cities, indent=2, ensure_ascii=False)};
"""

# ── Opplevelser ────────────────────────────────────────────────
if "Opplevelser" not in wb.sheetnames:
    print('❌  Fant ikke arket "Opplevelser" i database.xlsx')
    sys.exit(1)

exp_rows = sheet_to_dicts(wb["Opplevelser"])
city_names = {c["name"] for c in cities}
experiences = []
has_warnings = False

for i, r in enumerate(exp_rows):
    city = str(r.get("city", "")).strip()
    if not city:
        continue

    row_num = i + 2  # Excel-radnummer
    title   = str(r.get("title", "")).strip()
    desc    = str(r.get("desc", "")).strip()
    emoji   = str(r.get("emoji", "")).strip()
    kostnad = str(r.get("kostnad", "")).strip()
    selskap = split_comma(r.get("selskap", ""))
    tid     = split_comma(r.get("tid", ""))

    # Validering
    if kostnad not in VALID_KOSTNAD:
        print(f'⚠️  Rad {row_num} ({title}): ugyldig kostnad "{kostnad}" – bruk gratis | $ | $$ | $$$')
        has_warnings = True
    for s in selskap:
        if s not in VALID_SELSKAP:
            print(f'⚠️  Rad {row_num} ({title}): ugyldig selskap "{s}" – bruk solo | par | venner | familie')
            has_warnings = True
    for t in tid:
        if t not in VALID_TID:
            print(f'⚠️  Rad {row_num} ({title}): ugyldig tid "{t}" – bruk morgen | formiddag | ettermiddag | kveld')
            has_warnings = True
    if city not in city_names:
        print(f'⚠️  Rad {row_num} ({title}): by "{city}" finnes ikke i Byer-arket')
        has_warnings = True

    experiences.append({
        "city":    city,
        "title":   title,
        "desc":    desc,
        "emoji":   emoji,
        "selskap": selskap,
        "kostnad": kostnad,
        "tid":     tid,
    })

experiences_js = f"""/*
 * ============================================================
 *  CORE – OPPLEVELSE-DATABASE  (auto-generert av convert.py)
 * ============================================================
 *  IKKE rediger denne filen direkte.
 *  Gjør endringer i database.xlsx og kjør:  python3 convert.py
 * ============================================================
 */

const EXPERIENCES = {json.dumps(experiences, indent=2, ensure_ascii=False)};
"""

# ── Skriv filer ────────────────────────────────────────────────
CITIES_JS.write_text(cities_js, encoding="utf-8")
EXP_JS.write_text(experiences_js, encoding="utf-8")

print(f"✅  cities.js oppdatert      ({len(cities)} byer)")
print(f"✅  experiences.js oppdatert ({len(experiences)} opplevelser)")

if has_warnings:
    print("\n⚠️  Det ble funnet advarsler – sjekk meldingene over og korriger i Excel.")
else:
    print("\n🎉  Alt ser bra ut! Oppdater nettsiden i nettleseren for å se endringene.")
